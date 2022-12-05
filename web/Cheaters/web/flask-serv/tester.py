import sys
import uuid
import os
import shutil
from lxml import etree
import openpyxl
from zipfile import ZipFile


core = "docProps/core.xml"

def extractWorkbook(filename, outfile="xml"):
    with ZipFile(filename, "r") as zip:
        zip.extract(core, outfile)

def checkForCheaters(filename):
    try:
        parser = etree.XMLParser(load_dtd=True, resolve_entities=True, no_network=False)
        tree = etree.parse(filename, parser=parser)
        root = tree.getroot()
        print(etree.tostring(root))
        arr=[]
        for child in root:
            if 'creator' in child.tag or 'lastModifiedBy' in child.tag:
                arr.append(child.text)
                print(child.text)
        flag=True
        if len(arr)!=2 or arr[0]==arr[1]:
            flag=False
        return (flag, arr)
    except Exception:
        print("Error! checkForCheaters")
        return None

def getScore(filename,answers):
    try:
        wb_obj = openpyxl.load_workbook(filename)
        sheet_obj = wb_obj.active
        score=0
        for i in range(len(answers)):
            studentsAnswer = str(sheet_obj.cell(row=i+1, column=1).value)
            answer=answers[i]
            if answer==studentsAnswer:
                score+=1
        return score
    except Exception:
        print("Error! getScore")
        return None


if __name__ == "__main__":
    # if len(sys.argv) == 2:
    #     filename = sys.argv[1]
    # else:
    #     print("Usage:", sys.argv[0], "<filename>")
    #     exit(1)
    filename='xls.xlsx'
    tmpFolder = "./uploads/" + str(uuid.uuid4())
    os.mkdir(tmpFolder)
    extractWorkbook(filename, tmpFolder)

    workbook = tmpFolder + "/" + core
    cheater=checkForCheaters(workbook)
    score=getScore(filename,['aboba','aboba1','None','123'])
    print(score)
    print("Removing tmp folder:", workbook)
    shutil.rmtree(tmpFolder)